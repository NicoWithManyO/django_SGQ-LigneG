import os
import shutil
from datetime import datetime
from django.conf import settings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from production.models import Roll, CurrentProfile, Shift


class RollExcelExporter:
    """Service pour exporter les rouleaux dans un fichier Excel."""
    
    def __init__(self):
        self.excel_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        self.filename = 'rolls_export.xlsx'
        self.filepath = os.path.join(self.excel_dir, self.filename)
        self.max_rows = 10000  # Rotation après 10000 lignes
        
        # Créer le répertoire si nécessaire
        os.makedirs(self.excel_dir, exist_ok=True)
        
        # En-têtes des colonnes - TOUS les champs du modèle Roll
        self.headers = [
            'ID', 'ID Rouleau', 'OF', 'Numéro', 'Date/Heure', 'Opérateur', 'Shift',
            'Longueur (m)', 'Statut', 'Destination', 'Masse Tube (g)', 'Masse Totale (g)', 
            'Masse Nette (g)', 'Défauts Bloquants', 'Problèmes Épaisseur',
            'Épaisseur Moy. Gauche', 'Épaisseur Moy. Droite', 'Grammage Calc.',
            'Défauts', 'Profil', 'Commentaire'
        ]
    
    def _create_workbook(self):
        """Créer un nouveau fichier Excel avec les en-têtes."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Rouleaux"
        
        # Style pour les en-têtes
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # Ajouter les en-têtes
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            
        # Largeur des colonnes
        column_widths = {
            'A': 8,   # ID
            'B': 20,  # ID Rouleau
            'C': 10,  # OF
            'D': 10,  # Numéro
            'E': 20,  # Date/Heure
            'F': 20,  # Opérateur
            'G': 25,  # Shift
            'H': 12,  # Longueur (m)
            'I': 15,  # Statut
            'J': 15,  # Destination
            'K': 12,  # Masse Tube
            'L': 12,  # Masse Totale
            'M': 12,  # Masse Nette
            'N': 15,  # Défauts Bloquants
            'O': 18,  # Problèmes Épaisseur
            'P': 18,  # Épaisseur Moy. Gauche
            'Q': 18,  # Épaisseur Moy. Droite
            'R': 15,  # Grammage Calc.
            'S': 30,  # Défauts
            'T': 20,  # Profil
            'U': 30,  # Commentaire
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        return wb
    
    def _get_roll_data(self, roll):
        """Extraire TOUTES les données d'un rouleau pour l'export."""
        # Défauts
        defects = roll.defects.all()
        defects_str = ', '.join([
            f"{d.defect_type.name} à {d.meter_position}m" 
            for d in defects
        ])
        
        # Profil - récupérer depuis CurrentProfile ou depuis l'OF
        profile_name = ''
        
        # Essayer depuis CurrentProfile d'abord
        current_profile = CurrentProfile.objects.first()
        if current_profile and current_profile.profile:
            profile_name = current_profile.profile.name
        
        # Sinon essayer depuis l'OF
        if not profile_name and roll.fabrication_order and hasattr(roll.fabrication_order, 'profile'):
            profile_name = roll.fabrication_order.profile.name
        
        return [
            roll.id,  # ID unique de la base de données
            roll.roll_id,
            roll.fabrication_order.order_number if roll.fabrication_order else '',
            roll.roll_number or '',
            roll.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            f"{roll.shift.operator.first_name} {roll.shift.operator.last_name}" if roll.shift and roll.shift.operator else '',
            roll.shift.shift_id if roll.shift else (roll.shift_id_str or ''),
            float(roll.length) if roll.length else 0,
            roll.status or 'CONFORME',
            roll.destination or 'PRODUCTION',
            float(roll.tube_mass) if roll.tube_mass else 0,
            float(roll.total_mass) if roll.total_mass else 0,
            float(roll.net_mass) if roll.net_mass else 0,
            'Oui' if roll.has_blocking_defects else 'Non',
            'Oui' if roll.has_thickness_issues else 'Non',
            float(roll.avg_thickness_left) if roll.avg_thickness_left else '',
            float(roll.avg_thickness_right) if roll.avg_thickness_right else '',
            float(roll.grammage_calc) if roll.grammage_calc else '',
            defects_str,
            profile_name,
            roll.comment or ''
        ]
    
    def _check_rotation(self):
        """Vérifier si le fichier doit être archivé."""
        if os.path.exists(self.filepath):
            wb = load_workbook(self.filepath)
            ws = wb.active
            row_count = ws.max_row
            wb.close()
            
            if row_count >= self.max_rows:
                # Archiver le fichier
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                archive_name = f'rolls_export_{timestamp}.xlsx'
                archive_path = os.path.join(self.excel_dir, 'archives', archive_name)
                
                # Créer le répertoire d'archives
                os.makedirs(os.path.dirname(archive_path), exist_ok=True)
                
                # Déplacer le fichier
                shutil.move(self.filepath, archive_path)
                return True
        return False
    
    def export_roll(self, roll, update=True):
        """Ajouter ou mettre à jour un rouleau dans le fichier Excel."""
        try:
            # Vérifier la rotation
            self._check_rotation()
            
            # Charger ou créer le fichier
            if os.path.exists(self.filepath):
                wb = load_workbook(self.filepath)
                ws = wb.active
            else:
                wb = self._create_workbook()
                ws = wb.active
            
            # Récupérer les données du rouleau
            row_data = self._get_roll_data(roll)
            
            # Chercher si le rouleau existe déjà (par ID en colonne A)
            roll_row = None
            if update:
                for row_num in range(2, ws.max_row + 1):  # Commencer à 2 pour skip les en-têtes
                    if ws.cell(row=row_num, column=1).value == roll.id:
                        roll_row = row_num
                        break
            
            if roll_row:
                # Mettre à jour la ligne existante
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=roll_row, column=col, value=value)
            else:
                # Ajouter une nouvelle ligne
                ws.append(row_data)
                roll_row = ws.max_row
            
            # Style pour les rouleaux non conformes
            if roll.status != 'CONFORME':
                fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                for col in range(1, len(self.headers) + 1):
                    ws.cell(row=roll_row, column=col).fill = fill
            else:
                # Retirer le style rouge si le rouleau est devenu conforme
                for col in range(1, len(self.headers) + 1):
                    ws.cell(row=roll_row, column=col).fill = PatternFill()
            
            # Sauvegarder
            wb.save(self.filepath)
            wb.close()
            
            return True, self.filepath
            
        except Exception as e:
            return False, str(e)
    
    def delete_roll(self, roll):
        """Supprimer un rouleau du fichier Excel."""
        try:
            # Si le fichier n'existe pas, c'est un succès (rien à supprimer)
            if not os.path.exists(self.filepath):
                return True, "Fichier Excel inexistant"
            
            # Charger le fichier
            wb = load_workbook(self.filepath)
            ws = wb.active
            
            # Chercher la ligne par ID (colonne A)
            row_to_delete = None
            for row_num in range(2, ws.max_row + 1):  # Commencer à 2 pour skip les en-têtes
                if ws.cell(row=row_num, column=1).value == roll.id:
                    row_to_delete = row_num
                    break
            
            if row_to_delete:
                # Supprimer la ligne
                ws.delete_rows(row_to_delete, 1)
                
                # Sauvegarder
                wb.save(self.filepath)
                wb.close()
                
                return True, f"Ligne {row_to_delete} supprimée de {self.filepath}"
            else:
                # Rouleau non trouvé dans Excel, c'est OK
                wb.close()
                return True, "Rouleau non trouvé dans Excel"
                
        except Exception as e:
            return False, str(e)
    
    def get_export_path(self):
        """Retourner le chemin du fichier Excel."""
        return self.filepath


class ShiftExcelExporter:
    """Service pour exporter les shifts (postes) dans un fichier Excel."""
    
    def __init__(self):
        self.excel_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        self.filename = 'shifts_export.xlsx'
        self.filepath = os.path.join(self.excel_dir, self.filename)
        self.max_rows = 10000  # Rotation après 10000 lignes
        
        # Créer le répertoire si nécessaire
        os.makedirs(self.excel_dir, exist_ok=True)
        
        # En-têtes des colonnes - TOUS les champs du modèle Shift + TRS + agrégats
        self.headers = [
            'ID', 'Shift ID', 'Date', 'Vacation', 'Opérateur', 'Matricule',
            'Heure Début', 'Heure Fin', 'Temps Ouverture', 'Temps Disponible', 'Temps Perdu',
            'Longueur Totale (m)', 'Longueur Conforme (m)', 'Longueur Non-Conforme (m)', 
            'Déchet Brut (m)', 'Production Théorique (m)',
            'TRS (%)', 'Disponibilité (%)', 'Performance (%)', 'Qualité (%)',
            'Machine Début', 'Métrage Début', 'Machine Fin', 'Métrage Fin',
            'Épaisseur Moy. Gauche', 'Épaisseur Moy. Droite', 'Grammage Moyen',
            'Check-list Signée', 'Heure Signature', 'Nb Rouleaux Produits', 
            'Nb Rouleaux Conformes', 'Nb Temps Perdus', 'Profil Actuel', 'Vitesse Tapis',
            'Commentaires Opérateur', 'Créé le', 'Modifié le'
        ]
    
    def _create_workbook(self):
        """Créer un nouveau fichier Excel avec les en-têtes."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Shifts"
        
        # Style pour les en-têtes
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # Ajouter les en-têtes
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            
        # Largeur des colonnes optimisées
        column_widths = {
            'A': 8,   # ID
            'B': 25,  # Shift ID
            'C': 12,  # Date
            'D': 12,  # Vacation
            'E': 18,  # Opérateur
            'F': 15,  # Matricule
            'G': 12,  # Heure Début
            'H': 12,  # Heure Fin
            'I': 15,  # Temps Ouverture
            'J': 15,  # Temps Disponible
            'K': 15,  # Temps Perdu
            'L': 15,  # Longueur Totale
            'M': 18,  # Longueur Conforme
            'N': 20,  # Longueur Non-Conforme
            'O': 15,  # Déchet Brut
            'P': 18,  # Production Théorique
            'Q': 10,  # TRS
            'R': 12,  # Disponibilité
            'S': 12,  # Performance
            'T': 10,  # Qualité
            'U': 12,  # Machine Début
            'V': 15,  # Métrage Début
            'W': 12,  # Machine Fin
            'X': 15,  # Métrage Fin
            'Y': 18,  # Épaisseur Moy. Gauche
            'Z': 18,  # Épaisseur Moy. Droite
            'AA': 15, # Grammage Moyen
            'BB': 15, # Check-list Signée
            'CC': 15, # Heure Signature
            'DD': 18, # Nb Rouleaux Produits
            'EE': 18, # Nb Rouleaux Conformes
            'FF': 15, # Nb Temps Perdus
            'GG': 20, # Profil Actuel
            'HH': 15, # Vitesse Tapis
            'II': 30, # Commentaires Opérateur
            'JJ': 20, # Créé le
            'KK': 20, # Modifié le
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        return wb
    
    def _format_duration(self, duration):
        """Formater une durée en format lisible."""
        if not duration:
            return ''
        total_seconds = int(duration.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f"{hours:02d}:{minutes:02d}"
    
    def _get_shift_data(self, shift):
        """Extraire TOUTES les données d'un shift pour l'export."""
        
        # Données de base du shift
        operator_name = ''
        employee_id = ''
        if shift.operator:
            operator_name = f"{shift.operator.first_name} {shift.operator.last_name}"
            employee_id = shift.operator.employee_id
        
        # Données TRS (si disponibles)
        trs_data = {
            'opening_time': '',
            'availability_time': '',
            'lost_time': '',
            'total_length': 0,
            'ok_length': 0,
            'nok_length': 0,
            'raw_waste_length': 0,
            'theoretical_production': 0,
            'trs_percentage': 0,
            'availability_percentage': 0,
            'performance_percentage': 0,
            'quality_percentage': 0,
            'profile_name': '',
            'belt_speed': 0
        }
        
        if hasattr(shift, 'trs') and shift.trs:
            trs = shift.trs
            trs_data.update({
                'opening_time': self._format_duration(trs.opening_time),
                'availability_time': self._format_duration(trs.availability_time),
                'lost_time': self._format_duration(trs.lost_time),
                'total_length': float(trs.total_length) if trs.total_length else 0,
                'ok_length': float(trs.ok_length) if trs.ok_length else 0,
                'nok_length': float(trs.nok_length) if trs.nok_length else 0,
                'raw_waste_length': float(trs.raw_waste_length) if trs.raw_waste_length else 0,
                'theoretical_production': float(trs.theoretical_production) if trs.theoretical_production else 0,
                'trs_percentage': float(trs.trs_percentage) if trs.trs_percentage else 0,
                'availability_percentage': float(trs.availability_percentage) if trs.availability_percentage else 0,
                'performance_percentage': float(trs.performance_percentage) if trs.performance_percentage else 0,
                'quality_percentage': float(trs.quality_percentage) if trs.quality_percentage else 0,
                'profile_name': trs.profile_name or '',
                'belt_speed': float(trs.belt_speed_m_per_min) if trs.belt_speed_m_per_min else 0
            })
        
        # Agrégats depuis les rouleaux associés
        rolls = shift.rolls.all()
        nb_rolls_total = rolls.count()
        nb_rolls_conforme = rolls.filter(status='CONFORME').count()
        
        # Agrégats depuis les temps perdus
        from wcm.models import LostTimeEntry
        lost_times = LostTimeEntry.objects.filter(shift=shift)
        nb_lost_times = lost_times.count()
        
        # Formatage sécurisé des dates/heures
        from datetime import date, time, datetime
        
        def safe_date_format(value, format_str):
            """Formater une date/heure de manière sécurisée."""
            if not value:
                return ''
            if isinstance(value, str):
                return value  # Déjà formaté
            if hasattr(value, 'strftime'):
                return value.strftime(format_str)
            return str(value)
        
        return [
            shift.id,  # ID unique de la base de données
            shift.shift_id,
            safe_date_format(shift.date, '%Y-%m-%d'),
            shift.get_vacation_display() if shift.vacation else '',
            operator_name,
            employee_id,
            safe_date_format(shift.start_time, '%H:%M'),
            safe_date_format(shift.end_time, '%H:%M'),
            trs_data['opening_time'],
            trs_data['availability_time'],
            trs_data['lost_time'],
            trs_data['total_length'],
            trs_data['ok_length'],
            trs_data['nok_length'],
            trs_data['raw_waste_length'],
            trs_data['theoretical_production'],
            trs_data['trs_percentage'],
            trs_data['availability_percentage'],
            trs_data['performance_percentage'],
            trs_data['quality_percentage'],
            'Oui' if shift.started_at_beginning else 'Non',
            float(shift.meter_reading_start) if shift.meter_reading_start else '',
            'Oui' if shift.started_at_end else 'Non',
            float(shift.meter_reading_end) if shift.meter_reading_end else '',
            float(shift.avg_thickness_left_shift) if shift.avg_thickness_left_shift else '',
            float(shift.avg_thickness_right_shift) if shift.avg_thickness_right_shift else '',
            float(shift.avg_grammage_shift) if shift.avg_grammage_shift else '',
            shift.checklist_signed or '',
            safe_date_format(shift.checklist_signed_time, '%H:%M'),
            nb_rolls_total,
            nb_rolls_conforme,
            nb_lost_times,
            trs_data['profile_name'],
            trs_data['belt_speed'],
            shift.operator_comments or '',
            safe_date_format(shift.created_at, '%Y-%m-%d %H:%M:%S'),
            safe_date_format(shift.updated_at, '%Y-%m-%d %H:%M:%S')
        ]
    
    def _check_rotation(self):
        """Vérifier si le fichier doit être archivé."""
        if os.path.exists(self.filepath):
            wb = load_workbook(self.filepath)
            ws = wb.active
            row_count = ws.max_row
            wb.close()
            
            if row_count >= self.max_rows:
                # Archiver le fichier
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                archive_name = f'shifts_export_{timestamp}.xlsx'
                archive_path = os.path.join(self.excel_dir, 'archives', archive_name)
                
                # Créer le répertoire d'archives
                os.makedirs(os.path.dirname(archive_path), exist_ok=True)
                
                # Déplacer le fichier
                shutil.move(self.filepath, archive_path)
                return True
        return False
    
    def export_shift(self, shift, update=True):
        """Ajouter ou mettre à jour un shift dans le fichier Excel."""
        try:
            # Vérifier la rotation
            self._check_rotation()
            
            # Charger ou créer le fichier
            if os.path.exists(self.filepath):
                wb = load_workbook(self.filepath)
                ws = wb.active
            else:
                wb = self._create_workbook()
                ws = wb.active
            
            # Récupérer les données du shift
            row_data = self._get_shift_data(shift)
            
            # Chercher si le shift existe déjà (par ID en colonne A)
            shift_row = None
            if update:
                for row_num in range(2, ws.max_row + 1):  # Commencer à 2 pour skip les en-têtes
                    if ws.cell(row=row_num, column=1).value == shift.id:
                        shift_row = row_num
                        break
            
            if shift_row:
                # Mettre à jour la ligne existante
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=shift_row, column=col, value=value)
            else:
                # Ajouter une nouvelle ligne
                ws.append(row_data)
                shift_row = ws.max_row
            
            # Style conditionnel selon performance TRS
            if hasattr(shift, 'trs') and shift.trs:
                trs_percentage = float(shift.trs.trs_percentage) if shift.trs.trs_percentage else 0
                
                if trs_percentage < 60:  # TRS faible - rouge
                    fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif trs_percentage < 75:  # TRS moyen - orange
                    fill = PatternFill(start_color="FFEACC", end_color="FFEACC", fill_type="solid")
                elif trs_percentage >= 85:  # TRS excellent - vert
                    fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                else:  # TRS bon - pas de couleur
                    fill = PatternFill()
                
                for col in range(1, len(self.headers) + 1):
                    ws.cell(row=shift_row, column=col).fill = fill
            
            # Sauvegarder
            wb.save(self.filepath)
            wb.close()
            
            return True, self.filepath
            
        except Exception as e:
            return False, str(e)
    
    def delete_shift(self, shift):
        """Supprimer un shift du fichier Excel."""
        try:
            # Si le fichier n'existe pas, c'est un succès (rien à supprimer)
            if not os.path.exists(self.filepath):
                return True, "Fichier Excel inexistant"
            
            # Charger le fichier
            wb = load_workbook(self.filepath)
            ws = wb.active
            
            # Chercher la ligne par ID (colonne A)
            row_to_delete = None
            for row_num in range(2, ws.max_row + 1):  # Commencer à 2 pour skip les en-têtes
                if ws.cell(row=row_num, column=1).value == shift.id:
                    row_to_delete = row_num
                    break
            
            if row_to_delete:
                # Supprimer la ligne
                ws.delete_rows(row_to_delete, 1)
                
                # Sauvegarder
                wb.save(self.filepath)
                wb.close()
                
                return True, f"Ligne {row_to_delete} supprimée de {self.filepath}"
            else:
                # Shift non trouvé dans Excel, c'est OK
                wb.close()
                return True, "Shift non trouvé dans Excel"
                
        except Exception as e:
            return False, str(e)
    
    def get_export_path(self):
        """Retourner le chemin du fichier Excel."""
        return self.filepath